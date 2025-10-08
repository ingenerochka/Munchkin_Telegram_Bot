"""
Модуль для обработки загружаемого отчета по заявкам и конвертации заявок в ключи
"""

import logging

import psycopg
from openpyxl.reader.excel import load_workbook

import configs
import databases

def report(file_name):
    table = parse_excel(fr'downloads\{file_name}')
    table.pop(0)
    for engineer, task in table:
        keys = convert_to_keys(task)
        if keys > 0:
            update_keys(engineer, keys)
    return


# получаем данные из таблицы выполненных заявок
def parse_excel(file_path: str) -> list:
    """
    Функция принимает переменную, в которой указан путь к Excel-файлу, берет данные из определенных столбцов таблицы,
    записывает их в переменные и список
    :param file_path: строка с путем к файлу
    :return: список вида имя инженера:вид выполненной им заявки
    """
    data = []
    wb = load_workbook(filename=file_path)
    sheet = wb.active  # берём первый лист
    for row in sheet.iter_rows(values_only=True):  # пропускаем заголовки
        logging.debug(f"Активная строка отчета: {row}")
        engineer_name = row[12]  # берем ФИО инженера из 13-го столбца
        task_type = row[3]  # берем вид заявки из 4-го столбца
        data.append((engineer_name, task_type)) # в массив добавляются полученные данные о заявке
        if engineer_name == "None":
            data.pop()
    logging.info(f"Успешный парсинг заявок из Excel")
    return data

# переводим заявки в ключи
def convert_to_keys(task_type):
    """
    Функция принимает переменную, в которой указан вид заявки и сравнивает его со значениями словаря.
    Возвращает соответствующее виду заявки количество ключей или ноль, если соответствия не найдено.
    :param task_type: строка с видом выполненной заявки
    :return: кол-во ключей
    """
    logging.info(f"Успешный перевод вида заявки в ко-во ключей")
    return configs.task_types.get(task_type, 0)  # 0 если тип не найден


# запись данных о ключах в БД
def update_keys(engineer_name, keys_to_add):
    """
    Функция получает на вход ФИО инженера и кол-во ключей, которые он получил за выполненные заявки, подключается к БД, проверяет, есть ли инженер в БД.
    Если инженер есть, прибавляет полученные ключи к текущему значению ключей и записывает в БД. Если инженера нет, создает новую запись в БД с ФИО инженера и полученными ключами.
    :param engineer_name: ФИО инженера
    :param keys_to_add: количество ключей, полученных за выполненные заявки
    :return: None
    """
    # print(f'{engineer_name}, {keys_to_add}')
    # подключение к БД
    conn, cursor = databases.connet_to_db(configs.DB_CONFIG['database_main'])
    try:
        # Проверка, есть ли инженер в БД
        cursor.execute("SELECT keys FROM user_keys WHERE username = %s", (engineer_name,))
        result = cursor.fetchone()
        if result:
            new_keys = result[0] + keys_to_add
            cursor.execute("UPDATE user_keys SET keys = %s WHERE username = %s",
                           (new_keys, engineer_name))
            logging.info(f"Инженер {engineer_name} есть в БД. Ему успешно добавлено {keys_to_add} ключей, итого у него {new_keys} ключей.")
        else:
            cursor.execute("INSERT INTO user_keys (username, keys) VALUES (%s, %s)",
                           (engineer_name, keys_to_add))
            logging.info(f"Новый инженер добавлен в БД")
        conn.commit()
    except (Exception, psycopg.Error) as error:
        logging.exception(f"Ошибка добавления записи в БД: {error}")
    finally:
        cursor.close()
        conn.close()
